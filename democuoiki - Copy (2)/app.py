from flask import Flask, render_template, request,session, redirect, url_for,jsonify,request, send_file
import pandas as pd
import numpy as np
import os
from werkzeug.utils import secure_filename
import psycopg2
import datetime
import joblib
import io
import xlsxwriter
import json
import matplotlib
matplotlib.use('Agg') 
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import seaborn as sns
from weasyprint import HTML
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from flask import Flask, request, render_template, session, send_file
from reportlab.lib.styles import ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import base64
from reportlab.lib.pagesizes import A4, landscape

app = Flask(__name__)
app.secret_key = 'your_secret_key'  # Cần thiết để sử dụng session
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Tạo thư mục lưu biểu đồ nếu chưa có
if not os.path.exists('static/images'):
    os.makedirs('static/images')


# Đăng ký font RobotoCondensed-Regular
font_dir = os.path.join('static', 'fonts')
pdfmetrics.registerFont(TTFont('Roboto_Condensed-Regular', os.path.join(font_dir, 'Roboto_Condensed-Regular.ttf')))

models_path = os.path.join(os.getcwd(), 'models')

# Load các mô hình đã lưu
scaler = joblib.load(os.path.join(models_path, 'minmax_scaler.pkl'))
pca = joblib.load(os.path.join(models_path, 'pca_model.pkl'))

# Cấu hình matplotlib để sử dụng font hỗ trợ tiếng Việt
plt.rcParams['font.family'] = 'sans-serif'
plt.rcParams['font.sans-serif'] = ['Roboto_Condensed-Regular', 'Arial', 'DejaVu Sans']

@app.route("/", methods=["GET", "POST"])
def index():
    return render_template("index.html")

@app.route("/compare", methods=["GET", "POST"])
def compare():
    
    return render_template("compare.html")

# @app.route('/result', methods=['POST'])
# def result():
#     matrix_id = "matrix_" + datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

#     criteria = request.form.getlist('criteria[]')
#     option1_list = request.form.getlist('option1[]')  # có thể thêm các option khác nếu có

#     options = []
    
#     index = 1
#     while True:
#         key = f'option{index}[]'
#         if key in request.form:
#             values = [float(x) if x else 0.0 for x in request.form.getlist(key)]
#             options.append(values)
#             index += 1
#         else:
#             break

#     data_matrix = np.array(options).T
#     col_sums = np.sum(data_matrix, axis=0)
#     normalized_data = np.divide(data_matrix, col_sums, where=col_sums != 0)
#     rounded_data = np.round(normalized_data, 4)
#     weights = np.mean(rounded_data, axis=1)
#     weights = np.round(weights, 4).tolist()
#     Aw = np.dot(data_matrix, weights)
#     row_sums = np.sum(np.array([data_matrix[:, j] * weights[j] for j in range(len(weights))]), axis=0)
#     consistency_vector = np.divide(row_sums, weights, where=np.array(weights) != 0)
#     lambda_max = np.mean(consistency_vector)
#     n = len(criteria)
#     CI = (lambda_max - n) / (n - 1) if n > 1 else 0
#     RI_dict = {1: 0.0, 2: 0.0, 3: 0.58, 4: 0.90, 5: 1.12, 6: 1.24, 7: 1.32, 8: 1.41, 9: 1.45, 10: 1.49}
#     RI = RI_dict.get(n, 1.49)
#     CR = CI / RI if RI != 0 else 0


# # Tạo biểu đồ
#     # 1. Biểu đồ cột
#     plt.figure(figsize=(8, 5))
#     plt.bar(criteria, weights, color='skyblue')
#     plt.title('Trọng số của từng tiêu chí')
#     plt.xlabel('Tiêu chí')
#     plt.ylabel('Trọng số')
#     plt.grid(axis='y', linestyle='--', alpha=0.7)
#     # plt.savefig('static/images/bar_chart.png')
#     plt.close()


#     # 3. Biểu đồ nhiệt
#     plt.figure(figsize=(8, 6))
#     sns.heatmap(data_matrix, annot=True, fmt='.2f', cmap='Blues', xticklabels=criteria, yticklabels=criteria)
#     plt.title('Ma trận so sánh đôi')
#     # plt.savefig('static/images/heatmap.png')
#     plt.close()

#     # 👉 Lưu dữ liệu vào CSDL
#     try:
#         conn = get_connection()
#         cur = conn.cursor()

#         for j, col in enumerate(options):
#             for i, val in enumerate(col):
#                 cur.execute("""
#                     INSERT INTO CriteriaComparison (matrix_id, criteria_row, criteria_col, value)
#                     VALUES (%s, %s, %s, %s)
#                 """, (matrix_id, criteria[i], criteria[j], val))

#         conn.commit()
#         cur.close()
#         conn.close()
#     except Exception as e:
#         return f"❌ Đã xảy ra lỗi khi lưu vào CSDL: {str(e)}"

#     # Lưu vào session nếu cần sử dụng lại
#     session['criteria'] = criteria
#     session['weights'] = weights

#     return render_template(
#         'result.html',
#         criteria=criteria,
#         data=rounded_data.tolist(),
#         alternatives=criteria,
#         weights=weights,
#         aw=Aw.tolist(),
#         data_matrix=data_matrix.tolist(),
#         weighted_matrix=np.array([data_matrix[:, j] * weights[j] for j in range(len(weights))]).T.tolist(),
#         row_sums=row_sums.tolist(),
#         consistency_vector=consistency_vector.tolist(),
#         lambda_max=lambda_max,
#         ci=CI,
#         cr=CR,
#         weights_vector=weights
#     )
# ---------------------------------------------------dư đoan khi chưa cập nhật tính đối xứng-------------------------
# @app.route('/result', methods=['POST'])
# def result():
#     matrix_id = "matrix_" + datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

#     criteria = request.form.getlist('criteria[]')
#     option1_list = request.form.getlist('option1[]')

#     options = []
#     index = 1
#     while True:
#         key = f'option{index}[]'
#         if key in request.form:
#             values = [float(x) if x else 0.0 for x in request.form.getlist(key)]
#             options.append(values)
#             index += 1
#         else:
#             break

#     data_matrix = np.array(options).T
#     col_sums = np.sum(data_matrix, axis=0)
#     normalized_data = np.divide(data_matrix, col_sums, where=col_sums != 0)
#     rounded_data = np.round(normalized_data, 4)
#     weights = np.mean(rounded_data, axis=1)
#     weights = np.round(weights, 4).tolist()
#     Aw = np.dot(data_matrix, weights)
#     row_sums = np.sum(np.array([data_matrix[:, j] * weights[j] for j in range(len(weights))]), axis=0)
#     consistency_vector = np.divide(row_sums, weights, where=np.array(weights) != 0)
#     lambda_max = np.mean(consistency_vector)
#     n = len(criteria)
#     CI = (lambda_max - n) / (n - 1) if n > 1 else 0
#     RI_dict = {1: 0.0, 2: 0.0, 3: 0.58, 4: 0.90, 5: 1.12, 6: 1.24, 7: 1.32, 8: 1.41, 9: 1.45, 10: 1.49}
#     RI = RI_dict.get(n, 1.49)
#     CR = CI / RI if RI != 0 else 0

#     # Tạo biểu đồ
#     # 1. Biểu đồ cột
#     plt.figure(figsize=(8, 5))
#     plt.bar(criteria, weights, color='skyblue')
#     plt.title('Trọng số của từng tiêu chí')
#     plt.xlabel('Tiêu chí')
#     plt.ylabel('Trọng số')
#     plt.grid(axis='y', linestyle='--', alpha=0.7)
#     plt.savefig('static/images/bar_chart.png')
#     plt.close()

#     # 2. Biểu đồ radar
#     angles = np.linspace(0, 2 * np.pi, len(criteria), endpoint=False).tolist()
#     weights_plot = weights + weights[:1]
#     angles += angles[:1]
#     fig, ax = plt.subplots(figsize=(6, 6), subplot_kw=dict(polar=True))
#     ax.fill(angles, weights_plot, color='skyblue', alpha=0.25)
#     ax.plot(angles, weights_plot, color='blue', linewidth=2)
#     ax.set_yticklabels([])
#     ax.set_xticks(angles[:-1])
#     ax.set_xticklabels(criteria)
#     plt.title('Biểu đồ Radar - Trọng số tiêu chí')
#     plt.savefig('static/images/radar_chart.png')
#     plt.close()

#     # 3. Biểu đồ nhiệt
#     plt.figure(figsize=(8, 6))
#     sns.heatmap(data_matrix, annot=True, fmt='.2f', cmap='Blues', xticklabels=criteria, yticklabels=criteria)
#     plt.title('Ma trận so sánh đôi')
#     plt.savefig('static/images/heatmap.png')
#     plt.close()

#     # Lưu dữ liệu vào CSDL (giữ nguyên mã của bạn)
#     try:
#         conn = get_connection()
#         cur = conn.cursor()
#         for j, col in enumerate(options):
#             for i, val in enumerate(col):
#                 cur.execute("""
#                     INSERT INTO CriteriaComparison (matrix_id, criteria_row, criteria_col, value)
#                     VALUES (%s, %s, %s, %s)
#                 """, (matrix_id, criteria[i], criteria[j], val))
#         conn.commit()
#         cur.close()
#         conn.close()
#     except Exception as e:
#         return f"❌ Đã xảy ra lỗi khi lưu vào CSDL: {str(e)}"

#     session['criteria'] = criteria
#     session['weights'] = weights

#     return render_template(
#         'result.html',
#         criteria=criteria,
#         data=rounded_data.tolist(),
#         alternatives=criteria,
#         weights=weights,
#         aw=Aw.tolist(),
#         data_matrix=data_matrix.tolist(),
#         weighted_matrix=np.array([data_matrix[:, j] * weights[j] for j in range(len(weights))]).T.tolist(),
#         row_sums=row_sums.tolist(),
#         consistency_vector=consistency_vector.tolist(),
#         lambda_max=lambda_max,
#         ci=CI,
#         cr=CR,
#         weights_vector=weights
#     )


import unicodedata
@app.route('/result', methods=['POST'])
def result():
    matrix_id = "matrix_" + datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

    # Nhận dữ liệu từ form
    criteria = request.form.getlist('criteria[]')
# Chuẩn hóa ký tự Unicode
    criteria = [unicodedata.normalize('NFKD', c).encode('ascii', 'ignore').decode('ascii') for c in criteria]
    if not criteria or not all(isinstance(c, str) and c.strip() for c in criteria):
        print("Tiêu chí không hợp lệ từ form:", criteria)
        return jsonify({"error": "Tiêu chí không hợp lệ hoặc trống"}), 400

    options = []
    index = 1
    while True:
        key = f'option{index}[]'
        if key in request.form:
            values = [float(x) if x.strip() else 0.0 for x in request.form.getlist(key)]
            if len(values) != len(criteria):
                print(f"Dữ liệu phương án {index} không khớp:", values)
                return jsonify({"error": f"Dữ liệu phương án {index} không khớp với số tiêu chí"}), 400
            options.append(values)
            index += 1
        else:
            break

    if not options:
        print("Không có dữ liệu phương án")
        return jsonify({"error": "Không có dữ liệu phương án được cung cấp"}), 400

    # Tạo ma trận dữ liệu
    data_matrix = np.array(options).T
    col_sums = np.sum(data_matrix, axis=0)
    normalized_data = np.divide(data_matrix, col_sums, where=col_sums != 0)
    rounded_data = np.round(normalized_data, 4)
    weights = np.mean(rounded_data, axis=1)
    weights = np.round(weights, 4).tolist()

    # Tính toán AHP
    Aw = np.dot(data_matrix, weights)
    row_sums = np.sum(np.array([data_matrix[:, j] * weights[j] for j in range(len(weights))]), axis=0)
    consistency_vector = np.divide(row_sums, weights, where=np.array(weights) != 0)
    lambda_max = np.mean(consistency_vector)
    n = len(criteria)
    CI = (lambda_max - n) / (n - 1) if n > 1 else 0
    RI_dict = {1: 0.0, 2: 0.0, 3: 0.58, 4: 0.90, 5: 1.12, 6: 1.24, 7: 1.32, 8: 1.41, 9: 1.45, 10: 1.49}
    RI = RI_dict.get(n, 1.49)
    CR = CI / RI if RI != 0 else 0

    # Tạo biểu đồ
    plt.figure(figsize=(8, 5))
    plt.bar(criteria, weights, color='skyblue')
    plt.title('Trọng số của từng tiêu chí')
    plt.xlabel('Tiêu chí')
    plt.ylabel('Trọng số')
    plt.grid(axis='y', linestyle='--', alpha=0.7)
    plt.savefig('static/images/bar_chart.png')
    plt.close()

    angles = np.linspace(0, 2 * np.pi, len(criteria), endpoint=False).tolist()
    weights_plot = weights + weights[:1]
    angles += angles[:1]
    fig, ax = plt.subplots(figsize=(6, 6), subplot_kw=dict(polar=True))
    ax.fill(angles, weights_plot, color='skyblue', alpha=0.25)
    ax.plot(angles, weights_plot, color='blue', linewidth=2)
    ax.set_yticklabels([])
    ax.set_xticks(angles[:-1])
    ax.set_xticklabels(criteria)
    plt.title('Biểu đồ Radar - Trọng số tiêu chí')
    plt.savefig('static/images/radar_chart.png')
    plt.close()

    plt.figure(figsize=(8, 6))
    sns.heatmap(data_matrix, annot=True, fmt='.2f', cmap='Blues', xticklabels=criteria, yticklabels=[f'PA{i+1}' for i in range(len(options))])
    plt.title('Ma trận dữ liệu đầu vào')
    plt.savefig('static/images/heatmap.png')
    plt.close()

    # Lưu dữ liệu vào CSDL
    try:
        conn = get_connection()
        cur = conn.cursor()
        for j, col in enumerate(options):
            for i, val in enumerate(col):
                cur.execute("""
                    INSERT INTO CriteriaComparison (matrix_id, criteria_row, criteria_col, value)
                    VALUES (%s, %s, %s, %s)
                """, (matrix_id, criteria[i], f'PA{j+1}', val))
        conn.commit()
    except Exception as e:
        print(f"Lỗi lưu CSDL: {str(e)}")
        return jsonify({"error": f"Lỗi khi lưu vào CSDL: {str(e)}"}), 500
    finally:
        cur.close()
        conn.close()
# Kiểm tra JSON trước khi render
    try:
        criteria_json = json.dumps(criteria)
        weights_json = json.dumps(weights)
        print("JSON kiểm tra:", criteria_json, weights_json)
    except Exception as e:
        print(f"Lỗi tạo JSON trong result: {str(e)}")
        return jsonify({"error": f"Lỗi mã hóa JSON: {str(e)}"}), 500
    # Gỡ lỗi trước khi render
    print("Criteria trước khi render result.html:", criteria)
    print("Weights trước khi render result.html:", weights)

    return render_template(
        'result.html',
        criteria=criteria,
        data=rounded_data.tolist(),
        alternatives=[f'PA{i+1}' for i in range(len(options))],
        weights=weights,
        aw=Aw.tolist(),
        data_matrix=data_matrix.tolist(),
        weighted_matrix=np.array([data_matrix[:, j] * weights[j] for j in range(len(weights))]).T.tolist(),
        row_sums=row_sums.tolist(),
        consistency_vector=consistency_vector.tolist(),
        lambda_max=lambda_max,
        ci=CI,
        cr=CR,
        weights_vector=weights
    )



@app.route('/luachontieuchi', methods=['POST'])
def luachontieuchi():
    criteria = request.form.getlist('criteria[]')
    weights = request.form.getlist('weights[]')

    try:
        weights = [float(w) for w in weights]
    except ValueError:
        return "Lỗi: Dữ liệu trọng số không hợp lệ", 400

    # Debug (tùy chọn)
    print("Nhận được tiêu chí:", criteria)
    print("Nhận được trọng số:", weights)

    return render_template('luachontieuchi.html',
                           criteria=criteria,
                           weights=weights)


# @app.route('/luachontieuchi', methods=['GET', 'POST'])
# def luachontieuchi():
#     criteria = session.get('criteria', [])
#     weights = session.get('weights', [])

#     return render_template('luachontieuchi.html',
#                            criteria=criteria,
#                            weights=weights
#                            )


# -----------------------dự phong------------------------------------------------------------------

# @app.route('/combined_input', methods=['GET', 'POST'])
# def combined_input():
#     criteria = session.get('criteria', [])
#     weights = session.get('weights', [])

#     data = []
#     message = ""
#     scores = []  # Danh sách chứa điểm của từng phương án
#     best_option = None  # PA tốt nhất
#     # -------------Thêm__---------
#     # pairwise_matrices = {}
#     # eigen_vectors = {}
#     # lambda_max_list = []
#     # CI_list = []
#     # CR_list = []

#     # RI_dict = {1: 0.00, 2: 0.00, 3: 0.58, 4: 0.90, 5: 1.12, 6: 1.24, 7: 1.32, 8: 1.41, 9: 1.45, 10: 1.49}

#     if request.method == 'POST':
#         file = request.files.get('excel_file')

#         if file and file.filename.endswith('.xlsx'):
#             try:
#                 df = pd.read_excel(file)
#                 if len(df.columns) >= len(criteria):
#                     data = df.iloc[:3, :len(criteria)].values.tolist()
#                     message = "Đọc file Excel thành công!"
#             except Exception as e:
#                 message = f"Lỗi khi đọc file Excel: {e}"
#         else:
#             # Lấy dữ liệu nhập tay
#             for i in range(3):
#                 row = request.form.getlist(f'option{i+1}[]')
#                 if row:
#                     try:
#                         data.append([float(x) if x else 0.0 for x in row])
#                         # values = [float(x) if x and float(x) > 0 else 0.001 for x in row]
#                         # data.append(values)
#                     except:
#                         data.append([0.0 for _ in range(len(criteria))])
#             message = "Dữ liệu nhập tay đã được xử lý!"


#         # if data is not None and len(data) > 0 and len(data[0]) == len(criteria):
#         #     data = np.array(data)
#         #     data[data == 0] = 0.001

#         #     for j, crit in enumerate(criteria):
#         #         col = data[:, j]
#         #         matrix = np.zeros((3, 3))
#         #         for i in range(3):
#         #             for k in range(3):
#         #                 matrix[i][k] = col[i] / col[k]

#         #         pairwise_matrices[crit] = matrix
#         #         column_sums = np.sum(matrix, axis=0)
#         #         normalized_matrix = matrix / column_sums
#         #         eigen_vector = np.mean(normalized_matrix, axis=1)
#         #         eigen_vectors[crit] = eigen_vector.round(4)

#         #         lambda_max = np.sum(np.dot(matrix, eigen_vector) / eigen_vector) / 3
#         #         CI = (lambda_max - 3) / (3 - 1)
#         #         RI = RI_dict[3]
#         #         CR = CI / RI if RI != 0 else 0

#         #         lambda_max_list.append(round(lambda_max, 4))
#         #         CI_list.append(round(CI, 4))
#         #         CR_list.append(round(CR, 4))


#         # Tính độ ưu tiêu nếu có đủ dữ liệu và trọng số
#         try:
#             if data and weights and len(data[0]) == len(weights):
#                 for row in data:
#                     score = sum([v * w for v, w in zip(row, weights)])
#                     scores.append(round(score, 4))  # Làm tròn 4 chữ số thập phân
#                 best_option = scores.index(max(scores)) + 1  # PA tốt nhất (1-based index)
#         except Exception as e:
#             message += f" | Lỗi tính toán: {e}"

#     return render_template('combined_input.html',
#                            criteria=criteria,
#                            weights=weights,
#                            data=data,
#                            message=message,
#                            scores=scores,
#                            best_option=best_option,
#                         #    pairwise_matrices=pairwise_matrices,
#                         #    eigen_vectors=eigen_vectors,
#                         #    lambda_max_list=lambda_max_list,
#                         #    CI_list=CI_list,
#                         #    CR_list=CR_list 
#     )





# @app.route('/combined_input', methods=['GET', 'POST'])
# def combined_input():
#     # Tiêu chí mẫu (có thể thay bằng session nếu có từ form trước)
#     criteria = session.get('criteria', ['Chi phí', 'Chất lượng', 'Thời gian'])
#     weights = session.get('weights', [0.4, 0.4, 0.2])

#     data = []
#     message = ""
#     scores = []
#     best_option = None

#     pairwise_matrices = {}
#     eigen_vectors = {}
#     vectonhatquan = {}
#     lambda_max_list = []
#     CI_list = []
#     CR_list = []

#     charts = {}

#     RI_dict = {1: 0.0, 2: 0.0, 3: 0.58}

#     if request.method == 'POST':
#         # Nếu người dùng tải file Excel
#         file = request.files.get('excel_file')
#         if file and file.filename.endswith('.xlsx'):
#             try:
#                 df = pd.read_excel(file)
#                 data = df.iloc[:3, :len(criteria)].values.tolist()
#                 message = "Đọc dữ liệu từ Excel thành công!"
#             except Exception as e:
#                 message = f"Lỗi đọc file Excel: {e}"
#         else:
#             for i in range(3):
#                 row = request.form.getlist(f'option{i+1}[]')
#                 if row:
#                     try:
#                         row_values = [float(x) if x else 0.0 for x in row]
#                         data.append(row_values)
#                     except:
#                         data.append([0.0 for _ in range(len(criteria))])
#             message = "Đã xử lý dữ liệu nhập tay!"

#         if data and len(data[0]) == len(criteria):
#             for j, crit in enumerate(criteria):
#                 col = np.array([row[j] for row in data])
#                 matrix = np.zeros((3, 3))
#                 for i in range(3):
#                     for k in range(3):
#                         matrix[i][k] = ratio_to_ahp_scale(col[i], col[k])

#                 pairwise_matrices[crit] = matrix
#                 norm = matrix / np.sum(matrix, axis=0)
#                 eigen_vector = np.mean(norm, axis=1)
#                 eigen_vectors[crit] = eigen_vector.round(4)
#                 vectonhatquan[crit] = (np.dot(matrix, eigen_vector) / eigen_vector).round(4)

#                 lambda_max = np.mean(np.dot(matrix, eigen_vector) / eigen_vector)
#                 CI = (lambda_max - 3) / 2
#                 CR = CI / RI_dict[3] if RI_dict[3] != 0 else 0

#                 lambda_max_list.append(round(lambda_max, 4))
#                 CI_list.append(round(CI, 4))
#                 CR_list.append(round(CR, 4))

#                 save_pairwise_to_postgres(crit, matrix, eigen_vector)

#             # # Điểm tổng hợp
#             # for row in data:
#             #     score = sum(v * w for v, w in zip(row, weights))
#             #     scores.append(round(score, 4))
#             # best_option = scores.index(max(scores)) + 1


# #chuan hóa theo 1
# # Tính điểm tổng hợp ban đầu (chưa chuẩn hóa)
#             raw_scores = []
#             for row in data:
#                 score = sum(v * w for v, w in zip(row, weights))
#                 raw_scores.append(score)

#             # ✅ Chuẩn hóa để tổng = 1
#             total_score = sum(raw_scores)
#             scores = [round(score / total_score, 4) for score in raw_scores]

#             # Tìm phương án tốt nhất
#             best_option = scores.index(max(scores)) + 1
#             # Tạo các biểu đồ (matplotlib + seaborn)
#             def create_chart(fig):
#                 buf = io.BytesIO()
#                 fig.savefig(buf, format='png', bbox_inches='tight')
#                 buf.seek(0)
#                 return base64.b64encode(buf.getvalue()).decode('utf-8')

#             # Biểu đồ trọng số tiêu chí
#             fig1 = plt.figure()
#             plt.bar(criteria, weights, color='skyblue')
#             plt.title('Trọng số tiêu chí')
#             charts['criteria_weights_bar'] = create_chart(fig1)
#             plt.close()

#             # Biểu đồ điểm ưu tiên
#             fig2 = plt.figure()
#             plt.bar(['PA1', 'PA2', 'PA3'], scores, color='lightgreen')
#             plt.title('Điểm ưu tiên phương án')
#             charts['scores_bar'] = create_chart(fig2)
#             plt.close()

#             # Radar chart
#             angles = np.linspace(0, 2 * np.pi, len(criteria), endpoint=False).tolist()
#             angles += angles[:1]
#             fig3, ax = plt.subplots(subplot_kw={'polar': True})
#             for i, row in enumerate(data):
#                 values = row + row[:1]
#                 ax.plot(angles, values, label=f'PA{i+1}')
#                 ax.fill(angles, values, alpha=0.1)
#             ax.set_xticks(angles[:-1])
#             ax.set_xticklabels(criteria)
#             ax.set_title('Radar Chart')
#             ax.legend()
#             charts['options_radar'] = create_chart(fig3)
#             plt.close()

#             # Heatmap
#             fig4 = plt.figure()
#             sns.heatmap(pairwise_matrices[criteria[0]], annot=True, cmap='Blues', fmt='.2f',
#                         xticklabels=['PA1', 'PA2', 'PA3'], yticklabels=['PA1', 'PA2', 'PA3'])
#             plt.title(f'Ma trận so sánh cặp ({criteria[0]})')
#             charts['pairwise_heatmap_first'] = create_chart(fig4)
#             plt.close()

#             # Cột nhóm eigen vectors
#             fig5, ax = plt.subplots()
#             bar_width = 0.25
#             x = np.arange(len(criteria))
#             for i in range(3):
#                 values = [eigen_vectors[c][i] for c in criteria]
#                 ax.bar(x + i * bar_width, values, bar_width, label=f'PA{i+1}')
#             ax.set_xticks(x + bar_width)
#             ax.set_xticklabels(criteria)
#             ax.legend()
#             ax.set_title('Trọng số phương án theo tiêu chí')
#             charts['eigen_vectors_bar'] = create_chart(fig5)
#             plt.close()

#     return render_template('combined_input.html',
#                            criteria=criteria,
#                            weights=weights,
#                            data=data,
#                            message=message,
#                            scores=scores,
#                            best_option=best_option,
#                            pairwise_matrices={k: v.tolist() for k, v in pairwise_matrices.items()},
#                            eigen_vectors={k: v.tolist() for k, v in eigen_vectors.items()},
#                            vectonhatquan={k: v.tolist() for k, v in vectonhatquan.items()},
#                            lambda_max_list=lambda_max_list,
#                            CI_list=CI_list,
#                            CR_list=CR_list,
#                            charts=charts)
# phần này mới có thể lấy lại


# @app.route('/combined_input', methods=['GET', 'POST'])
# def combined_input():
#     # Tiêu chí mẫu (có thể thay bằng session nếu có từ form trước)
#     criteria = session.get('criteria', ['Pin', 'Hiệu Suất', 'Camera', 'Màn Hình', 'Trọng Lượng', 'Kết Nối', 'Bộ Nhớ'])
#     weights = session.get('weights', [0.3622, 0.2569, 0.1458, 0.1006, 0.063, 0.0431, 0.0283])

#     data = []
#     message = ""
#     scores = []
#     best_option = None

#     pairwise_matrices = {}
#     eigen_vectors = {}
#     vectonhatquan = {}
#     lambda_max_list = []
#     CI_list = []
#     CR_list = []

#     charts = {}

#     RI_dict = {1: 0.0, 2: 0.0, 3: 0.58}

#     if request.method == 'POST':
#         # Nếu người dùng tải file Excel
#         file = request.files.get('excel_file')
#         if file and file.filename.endswith('.xlsx'):
#             try:
#                 df = pd.read_excel(file)
#                 data = df.iloc[:3, :len(criteria)].values.tolist()
#                 print("Dữ liệu nhập tay:", data)  # In dữ liệu để kiểm tra
#                 message = "Đọc dữ liệu từ Excel thành công!"
#             except Exception as e:
#                 message = f"Lỗi đọc file Excel: {e}"
#         else:
#             # Xử lý dữ liệu nhập tay
#             for i in range(3):
#                 row = request.form.getlist(f'option{i+1}[]')
#                 if row:
#                     try:
#                         row_values = [float(x) if x else 0.0 for x in row]
#                         data.append(row_values)
#                         print("Dữ liệu nhập tay:", data)  # In dữ liệu để kiểm tra
#                     except:
#                         data.append([0.0 for _ in range(len(criteria))])
#             message = "Đã xử lý dữ liệu nhập tay!"

#         if data and len(data[0]) == len(criteria):
#             for j, crit in enumerate(criteria):
#                 col = np.array([row[j] for row in data])
#                 matrix = np.zeros((3, 3))
#                 for i in range(3):
#                     for k in range(3):
#                         matrix[i][k] = ratio_to_ahp_scale(col[i], col[k])

#                 pairwise_matrices[crit] = matrix
#                 norm = matrix / np.sum(matrix, axis=0)
#                 eigen_vector = np.mean(norm, axis=1)
#                 eigen_vectors[crit] = eigen_vector.round(4)
#                 vectonhatquan[crit] = (np.dot(matrix, eigen_vector) / eigen_vector).round(4)

#                 lambda_max = np.mean(np.dot(matrix, eigen_vector) / eigen_vector)
#                 CI = (lambda_max - 3) / 2
#                 CR = CI / RI_dict[3] if RI_dict[3] != 0 else 0

#                 lambda_max_list.append(round(lambda_max, 4))
#                 CI_list.append(round(CI, 4))
#                 CR_list.append(round(CR, 4))

#                 # Lưu ma trận vào cơ sở dữ liệu (nếu cần)
#                 save_pairwise_to_postgres(crit, matrix, eigen_vector)

#             # Tính điểm tổng hợp
#             raw_scores = []
#             for row in data:
#                 score = sum(v * w for v, w in zip(row, weights))
#                 raw_scores.append(score)

#             # Chuẩn hóa để tổng = 1
#             total_score = sum(raw_scores)
#             scores = [round(score / total_score, 4) for score in raw_scores]
#             best_option = scores.index(max(scores)) + 1

#             # Lưu tất cả dữ liệu vào session
#             session['criteria'] = criteria
#             session['weights'] = weights
#             session['data'] = data
#             print("Received data:", data)
#             session['pairwise_matrices'] = {k: v.tolist() for k, v in pairwise_matrices.items()}
#             session['eigen_vectors'] = {k: v.tolist() for k, v in eigen_vectors.items()}
#             session['vectonhatquan'] = {k: v.tolist() for k, v in vectonhatquan.items()}
#             session['lambda_max_list'] = lambda_max_list
#             session['CI_list'] = CI_list
#             session['CR_list'] = CR_list
#             session['scores'] = scores
#             session['best_option'] = best_option

#             # Tạo các biểu đồ (matplotlib + seaborn)
#             def create_chart(fig):
#                 buf = io.BytesIO()
#                 fig.savefig(buf, format='png', bbox_inches='tight')
#                 buf.seek(0)
#                 return base64.b64encode(buf.getvalue()).decode('utf-8')

#             # Biểu đồ trọng số tiêu chí
#             fig1 = plt.figure()
#             plt.bar(criteria, weights, color='skyblue')
#             plt.title('Trọng số tiêu chí')
#             charts['criteria_weights_bar'] = create_chart(fig1)
#             plt.close()

#             # Biểu đồ điểm ưu tiên
#             fig2 = plt.figure()
#             plt.bar(['PA1', 'PA2', 'PA3'], scores, color='lightgreen')
#             plt.title('Điểm ưu tiên phương án')
#             charts['scores_bar'] = create_chart(fig2)
#             plt.close()

#             # Radar chart
#             angles = np.linspace(0, 2 * np.pi, len(criteria), endpoint=False).tolist()
#             angles += angles[:1]
#             fig3, ax = plt.subplots(subplot_kw={'polar': True})
#             for i, row in enumerate(data):
#                 values = row + row[:1]
#                 ax.plot(angles, values, label=f'PA{i+1}')
#                 ax.fill(angles, values, alpha=0.1)
#             ax.set_xticks(angles[:-1])
#             ax.set_xticklabels(criteria)
#             ax.set_title('Radar Chart')
#             ax.legend()
#             charts['options_radar'] = create_chart(fig3)
#             plt.close()

#             # Heatmap
#             fig4 = plt.figure()
#             sns.heatmap(pairwise_matrices[criteria[0]], annot=True, cmap='Blues', fmt='.2f',
#                         xticklabels=['PA1', 'PA2', 'PA3'], yticklabels=['PA1', 'PA2', 'PA3'])
#             plt.title(f'Ma trận so sánh cặp ({criteria[0]})')
#             charts['pairwise_heatmap_first'] = create_chart(fig4)
#             plt.close()

#             # Cột nhóm eigen vectors
#             fig5, ax = plt.subplots()
#             bar_width = 0.25
#             x = np.arange(len(criteria))
#             for i in range(3):
#                 values = [eigen_vectors[c][i] for c in criteria]
#                 ax.bar(x + i * bar_width, values, bar_width, label=f'PA{i+1}')
#             ax.set_xticks(x + bar_width)
#             ax.set_xticklabels(criteria)
#             ax.legend()
#             ax.set_title('Trọng số phương án theo tiêu chí')
#             charts['eigen_vectors_bar'] = create_chart(fig5)
#             plt.close()

#             # Lưu charts vào session
#             session['charts'] = charts

#     return render_template('combined_input.html',
#                            criteria=criteria,
#                             weights=weights,
#                             data=data,
#                             message=message,
#                             scores=scores,
#                             best_option=best_option,
#                             pairwise_matrices={k: v.tolist() for k, v in pairwise_matrices.items()},
#                             eigen_vectors={k: v.tolist() for k, v in eigen_vectors.items()},
#                             vectonhatquan={k: v.tolist() for k, v in vectonhatquan.items()},
#                             lambda_max_list=lambda_max_list,
#                             CI_list=CI_list,
#                             CR_list=CR_list,
#                             charts=charts)

def convert_numpy_types(obj):
    """Chuyển đổi các kiểu NumPy thành kiểu Python native."""
    if isinstance(obj, np.integer):
        return int(obj)
    elif isinstance(obj, np.floating):
        return float(obj)
    elif isinstance(obj, np.ndarray):
        return obj.tolist()
    elif isinstance(obj, list):
        return [convert_numpy_types(item) for item in obj]
    elif isinstance(obj, dict):
        return {k: convert_numpy_types(v) for k, v in obj.items()}
    return obj

@app.route('/combined_input', methods=['GET', 'POST'])
def combined_input():
    criteria = []
    weights = []
    data = []
    message = None
    error = None
    scores = None
    best_option = None
    pairwise_matrices = {}
    eigen_vectors = {}
    vectonhatquan = {}
    lambda_max_list = []
    CI_list = []
    CR_list = []
    charts = {}

    if request.method == 'POST':
        try:
            # Nhận dữ liệu từ form
            criteria = request.form.getlist('criteria[]')
            weights_raw = request.form.getlist('weights[]')
            criteria = [unicodedata.normalize('NFKD', c).encode('ascii', 'ignore').decode('utf-8') for c in criteria]

            # Gỡ lỗi
            print("Dữ liệu nhận được trong combined_input:")
            print(f"criteria: {criteria}")
            print(f"weights_raw: {weights_raw}")

            # Kiểm tra criteria
            if not criteria or not all(isinstance(c, str) and c.strip() for c in criteria):
                error = "Tiêu chí không hợp lệ hoặc trống"
                return render_template('combined_input.html', error=error, criteria=criteria, weights=weights, data=data)

            # Chuyển đổi weights
            try:
                weights = [float(w) for w in weights_raw]
            except ValueError:
                error = "Trọng số chứa giá trị không hợp lệ"
                return render_template('combined_input.html', error=error, criteria=criteria, weights=weights, data=data)

            if len(weights) != len(criteria) or not all(isinstance(w, (int, float)) for w in weights):
                error = "Trọng số không hợp lệ hoặc không khớp với số tiêu chí"
                return render_template('combined_input.html', error=error, criteria=criteria, weights=weights, data=data)

            # Nhận dữ liệu từ form hoặc file Excel
            if 'excel_file' in request.files and request.files['excel_file'].filename:
                # Xử lý file Excel
                file = request.files['excel_file']
                if not file.filename.endswith('.xlsx'):
                    error = "File phải là định dạng .xlsx"
                    return render_template('combined_input.html', error=error, criteria=criteria, weights=weights, data=data)
                try:
                    df = pd.read_excel(file, engine='openpyxl')
                    if list(df.columns) != criteria:
                        error = f"Các cột trong Excel phải khớp với tiêu chí: {criteria}"
                        return render_template('combined_input.html', error=error, criteria=criteria, weights=weights, data=data)
                    data = df.values.tolist()
                    data = convert_numpy_types(data)  # Chuyển đổi kiểu NumPy
                    message = "Dữ liệu đã được tải từ file Excel."
                except Exception as e:
                    error = f"Lỗi đọc file Excel: {str(e)}"
                    return render_template('combined_input.html', error=error, criteria=criteria, weights=weights, data=data)
            else:
                # Nhận dữ liệu từ form (nhập tay)
                data = []
                for i in range(3):  # PA1, PA2, PA3
                    option = request.form.getlist(f'option{i+1}[]')
                    if len(option) != len(criteria):
                        error = f"Dữ liệu phương án PA{i+1} không khớp với số tiêu chí"
                        return render_template('combined_input.html', error=error, criteria=criteria, weights=weights, data=data)
                    try:
                        data.append([float(x) if x.strip() else 0.0 for x in option])
                    except ValueError:
                        error = f"Dữ liệu phương án PA{i+1} chứa giá trị không hợp lệ"
                        return render_template('combined_input.html', error=error, criteria=criteria, weights=weights, data=data)
                message = "Dữ liệu đã được nhập tay thành công."

            # Kiểm tra dữ liệu
            if not data or len(data) != 3:
                error = "Phải cung cấp dữ liệu cho 3 phương án"
                return render_template('combined_input.html', error=error, criteria=criteria, weights=weights, data=data)

            # Kiểm tra giá trị 0
            if any(data[i][j] == 0 for i in range(3) for j in range(len(criteria))):
                print("Cảnh báo: Dữ liệu chứa giá trị 0, thay bằng 0.001 để tránh chia cho 0")
                data = [[max(x, 0.001) for x in row] for row in data]

            # Tạo ma trận so sánh cặp với thang AHP
            for crit_idx, crit in enumerate(criteria):
                matrix = np.zeros((3, 3))
                for i in range(3):
                    for j in range(3):
                        if i == j:
                            matrix[i][j] = 1.0
                        elif i < j:  # Chỉ tính nửa trên
                            val_i = data[i][crit_idx]
                            val_j = data[j][crit_idx]
                            matrix[i][j] = ratio_to_ahp_scale(val_i, val_j)
                            matrix[j][i] = 1.0 / matrix[i][j]  # Đảm bảo đối xứng
                pairwise_matrices[crit] = convert_numpy_types(matrix)

                # Tính eigenvector
                eigenvalues, eigenvectors = np.linalg.eig(matrix)
                max_eigenvalue = np.max(np.real(eigenvalues))
                max_eigenvector = np.real(eigenvectors[:, np.argmax(np.real(eigenvalues))])
                max_eigenvector = max_eigenvector / np.sum(max_eigenvector)
                eigen_vectors[crit] = convert_numpy_types(max_eigenvector)

                # Tính vector nhất quán
                Aw = np.dot(matrix, max_eigenvector)
                consistency_vector = Aw / max_eigenvector
                vectonhatquan[crit] = convert_numpy_types(consistency_vector)

                # Tính λ max, CI, CR
                lambda_max = np.mean(consistency_vector)
                n = 3
                CI = (lambda_max - n) / (n - 1) if n > 1 else 0
                RI = 0.58  # RI cho n=3
                CR = CI / RI if RI != 0 else 0

                lambda_max_list.append(round(float(lambda_max), 4))
                CI_list.append(round(float(CI), 4))
                CR_list.append(round(float(CR), 4))

            # Tính điểm ưu tiên
            scores = np.zeros(3)
            for crit_idx, crit in enumerate(criteria):
                for i in range(3):
                    scores[i] += eigen_vectors[crit][i] * weights[crit_idx]
            scores = convert_numpy_types(scores)

            # Xác định phương án tốt nhất
            best_option = int(np.argmax(scores)) + 1

            # Tạo biểu đồ
            # Biểu đồ cột - Trọng số tiêu chí
            fig, ax = plt.subplots(figsize=(8, 5))
            ax.bar(criteria, weights, color='skyblue')
            ax.set_title('Trọng số của từng tiêu chí')
            ax.set_xlabel('Tiêu chí')
            ax.set_ylabel('Trọng số')
            ax.grid(axis='y', linestyle='--', alpha=0.7)
            plt.xticks(rotation=45, ha='right')
            buf = io.BytesIO()
            plt.savefig(buf, format='png', bbox_inches='tight')
            buf.seek(0)
            charts['criteria_weights_bar'] = base64.b64encode(buf.getvalue()).decode('utf-8')
            plt.close()

            # Biểu đồ cột - Điểm ưu tiên
            fig, ax = plt.subplots(figsize=(8, 5))
            ax.bar(['PA1', 'PA2', 'PA3'], scores, color='lightgreen')
            ax.set_title('Điểm ưu tiên của các phương án')
            ax.set_xlabel('Phương án')
            ax.set_ylabel('Điểm ưu tiên')
            ax.grid(axis='y', linestyle='--', alpha=0.7)
            buf = io.BytesIO()
            plt.savefig(buf, format='png')
            buf.seek(0)
            charts['scores_bar'] = base64.b64encode(buf.getvalue()).decode('utf-8')
            plt.close()

            # Biểu đồ radar
            angles = np.linspace(0, 2 * np.pi, len(criteria), endpoint=False).tolist()
            angles += angles[:1]
            fig, ax = plt.subplots(figsize=(6, 6), subplot_kw=dict(polar=True))
            for i in range(3):
                values = [eigen_vectors[crit][i] for crit in criteria] + [eigen_vectors[criteria[0]][i]]
                ax.plot(angles, values, label=f'PA{i+1}')
                ax.fill(angles, values, alpha=0.1)
            ax.set_xticks(angles[:-1])
            ax.set_xticklabels(criteria)
            ax.set_title('Biểu đồ Radar - Điểm số phương án')
            ax.legend()
            buf = io.BytesIO()
            plt.savefig(buf, format='png')
            buf.seek(0)
            charts['options_radar'] = base64.b64encode(buf.getvalue()).decode('utf-8')
            plt.close()

            # Biểu đồ nhiệt - Ma trận so sánh cặp (tiêu chí đầu tiên)
            fig, ax = plt.subplots(figsize=(8, 6))
            sns.heatmap(pairwise_matrices[criteria[0]], annot=True, fmt='.2f', cmap='Blues', ax=ax,
                        xticklabels=['PA1', 'PA2', 'PA3'], yticklabels=['PA1', 'PA2', 'PA3'])
            ax.set_title(f'Ma trận so sánh cặp - Tiêu chí: {criteria[0]}')
            buf = io.BytesIO()
            plt.savefig(buf, format='png')
            buf.seek(0)
            charts['pairwise_heatmap_first'] = base64.b64encode(buf.getvalue()).decode('utf-8')
            plt.close()

            # Biểu đồ cột nhóm - Trọng số phương án theo tiêu chí
            fig, ax = plt.subplots(figsize=(10, 6))
            x = np.arange(len(criteria))
            width = 0.25
            for i in range(3):
                ax.bar(x + i * width, [eigen_vectors[crit][i] for crit in criteria], width, label=f'PA{i+1}')
            ax.set_xticks(x + width)
            ax.set_xticklabels(criteria)
            ax.set_title('Trọng số phương án theo tiêu chí')
            ax.set_ylabel('Trọng số')
            ax.legend()
            ax.grid(axis='y', linestyle='--', alpha=0.7)
            plt.xticks(rotation=45, ha='right')
            buf = io.BytesIO()
            plt.savefig(buf, format='png', bbox_inches='tight')
            buf.seek(0)
            charts['eigen_vectors_bar'] = base64.b64encode(buf.getvalue()).decode('utf-8')
            plt.close()

        except Exception as e:
            print(f"Lỗi trong combined_input: {str(e)}")
            error = f"Lỗi xử lý dữ liệu: {str(e)}"
            return render_template('combined_input.html', error=error, criteria=criteria, weights=weights, data=data)


    # Render template
    return render_template('combined_input.html',
                           criteria=criteria,
                           weights=weights,
                           data=data,
                           message=message,
                           error=error,
                           scores=scores,
                           best_option=best_option,
                           pairwise_matrices=pairwise_matrices,
                           eigen_vectors=eigen_vectors,
                           vectonhatquan=vectonhatquan,
                           lambda_max_list=lambda_max_list,
                           CI_list=CI_list,
                           CR_list=CR_list,
                           charts=charts)





def generate_matrix_id():
    # Lấy thời gian hiện tại và định dạng thành YYYYMMDD_HHMMSS
    # current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
    current_time = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"matrix_{current_time}"

# # Hàm quy đổi tỉ lệ sang thang AHP
# def ratio_to_ahp_scale(val1, val2):
#     if val2 == 0:
#         return 1
#     ratio = val1 / val2
#     if ratio >= 9:
#         return 9
#     elif ratio <= 1/9:
#         return 1/9
#     else:
#         return round(ratio, 2)
# Hàm quy đổi tỉ lệ sang thang chuẩn AHP (1/9 đến 9)
def ratio_to_ahp_scale(val1, val2):
    ahp_scale = [1/9, 1/8, 1/7, 1/6, 1/5, 1/4, 1/3, 1/2,
                 1, 2, 3, 4, 5, 6, 7, 8, 9]
    
    if val2 == 0:
        return 1  # hoặc raise lỗi

    ratio = val1 / val2

    # Tìm giá trị gần nhất trong thang AHP
    closest = min(ahp_scale, key=lambda x: abs(x - ratio))
    return closest



# Hàm lưu dữ liệu cácác phương án vào PostgreSQL
def save_to_postgres(data, weights):
    conn = get_connection()
    cur = conn.cursor()

    # Xóa dữ liệu cũ (tùy chọn, nếu muốn giữ lịch sử thì bỏ dòng này)
    cur.execute("DELETE FROM options_data")

    # Tạo matrix_id cho nhóm bản ghi
    matrix_id = generate_matrix_id()

    # Lưu dữ liệu phương án
    for i, row in enumerate(data, 1):
        cur.execute(
            "INSERT INTO options_data (option_id, pin, performance, camera, screen, weight, connectivity, memory, matrix_id) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)",
            (f'PA{i}', row[0], row[1], row[2], row[3], row[4], row[5], row[6], matrix_id)
        )

    # Lưu trọng số
    cur.execute(
        "INSERT INTO options_data (option_id, pin, performance, camera, screen, weight, connectivity, memory, matrix_id) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)",
        ('Weights', weights[0], weights[1], weights[2], weights[3], weights[4], weights[5], weights[6], matrix_id)
    )

    conn.commit()
    cur.close()
    conn.close()



# Hàm lưu ma trận so sánh cặp và trọng số vào PostgreSQL
def save_pairwise_to_postgres(criterion, matrix, weights):
    conn = get_connection()
    cur = conn.cursor()

    # Xóa dữ liệu cũ cho tiêu chí này (nếu cần)
    cur.execute("DELETE FROM pairwise_comparison WHERE criterion = %s", (criterion,))

    # Tạo matrix_id cho nhóm bản ghi
    matrix_id = generate_matrix_id()

    # Lưu ma trận so sánh cặp
    for i in range(3):
        for j in range(3):
            cur.execute(
                "INSERT INTO pairwise_comparison (criterion, option1, option2, pairwise_value, weight, matrix_id) VALUES (%s, %s, %s, %s, %s, %s)",
                (criterion, f'PA{i+1}', f'PA{j+1}', float(matrix[i][j]), None, matrix_id)
            )

    # Lưu trọng số
    for i, weight in enumerate(weights, 1):
        cur.execute(
            "INSERT INTO pairwise_comparison (criterion, option1, option2, pairwise_value, weight, matrix_id) VALUES (%s, %s, %s, %s, %s, %s)",
            (criterion, f'PA{i}', 'Weight', None, float(weight), matrix_id)
        )

    conn.commit()
    cur.close()
    conn.close()


# Hàm lưu điểm ưu tiên vào PostgreSQL
def save_scores(scores):
    conn = get_connection()
    cur = conn.cursor()

    # Xóa dữ liệu cũ (nếu cần)
    cur.execute("DELETE FROM option_scores")

    # Tạo matrix_id cho nhóm bản ghi
    matrix_id = generate_matrix_id()

    # Lưu điểm ưu tiên
    for i, score in enumerate(scores, 1):
        cur.execute(
            "INSERT INTO option_scores (option_id, score, matrix_id) VALUES (%s, %s, %s)",
            (f'PA{i}', float(score), matrix_id)
        )

    conn.commit()
    cur.close()
    conn.close()

def convert_numpy_types(obj):
    """Chuyển đổi các kiểu NumPy thành kiểu Python native."""
    if isinstance(obj, np.integer):
        return int(obj)
    elif isinstance(obj, np.floating):
        return float(obj)
    elif isinstance(obj, np.ndarray):
        return obj.tolist()
    elif isinstance(obj, list):
        return [convert_numpy_types(item) for item in obj]
    elif isinstance(obj, dict):
        return {k: convert_numpy_types(v) for k, v in obj.items()}
    return obj



import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
@app.route('/export_excel', methods=['POST'])

def export_excel():
    print("[DEBUG] Dữ liệu POST:", request.form)

    try:
        # Helper function
        def load_json_field(key):
            val = request.form.get(key, '')
            if not val:
                raise ValueError(f"Thiếu hoặc rỗng: {key}")
            return json.loads(val)

        # Load form data
        data = load_json_field('data')
        criteria = load_json_field('criteria')
        weights = load_json_field('weights')
        pairwise_matrices = load_json_field('pairwise_matrices')
        eigen_vectors = load_json_field('eigen_vectors')
        vectonhatquan = load_json_field('vectonhatquan')
        lambda_max_list = load_json_field('lambda_max_list')
        CI_list = load_json_field('CI_list')
        CR_list = load_json_field('CR_list')
        scores = load_json_field('scores')
        best_option = load_json_field('best_option')

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tổng quan AHP"

        bold_center = Font(bold=True)
        center = Alignment(horizontal="center")

        # Sheet 1: Ma trận phương án
        ws.append(["Phương án"] + criteria)
        for i, row in enumerate(data):
            ws.append([f"PA{i+1}"] + row)
        ws.append(["Trọng số"] + weights)

        # Auto-width
        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

        # Format headers
        for cell in ws[1]:
            cell.font = bold_center
            cell.alignment = center

        # Sheet 2+: Ma trận so sánh cặp từng tiêu chí
        for idx, crit in enumerate(criteria):
            ws_matrix = wb.create_sheet(title=f"So sánh - {crit}")
            ws_matrix.append([""] + [f"PA{i+1}" for i in range(3)] + ["Trọng số", "Vecto nhất quán"])
            for i in range(3):
                row = [f"PA{i+1}"]
                row += pairwise_matrices[crit][i]
                row += [eigen_vectors[crit][i], vectonhatquan[crit][i]]
                ws_matrix.append(row)

            ws_matrix.append([])
            ws_matrix.append(["λ max", lambda_max_list[idx]])
            ws_matrix.append(["CI", CI_list[idx]])
            ws_matrix.append(["CR", CR_list[idx]])

            for cell in ws_matrix[1]:
                cell.font = bold_center
                cell.alignment = center
            for col in ws_matrix.columns:
                max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
                ws_matrix.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

        # Sheet cuối: Tổng hợp điểm ưu tiên
        ws_result = wb.create_sheet(title="Kết quả")
        ws_result.append(["Phương án", "Điểm ưu tiên"])
        for i, score in enumerate(scores):
            ws_result.append([f"PA{i+1}", score])
        ws_result.append([])
        ws_result.append(["Phương án tốt nhất", f"PA{best_option}"])

        for cell in ws_result[1]:
            cell.font = bold_center
            cell.alignment = center

        # Xuất file
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output,
                         as_attachment=True,
                         download_name="AHP_Results.xlsx",
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        print(f"Lỗi trong export_excel: {e}")
        return f"Lỗi xuất Excel: {str(e)}", 500


# Đăng ký font Roboto hoặc Arial có hỗ trợ Unicode
pdfmetrics.registerFont(TTFont('Roboto', 'static/fonts/Roboto_Condensed-Black.ttf'))
@app.route('/export_pdf', methods=['POST', 'GET'])
def export_pdf():
    styles = getSampleStyleSheet()

        # Tạo styles mới hỗ trợ tiếng Việt
    styleH = ParagraphStyle(name='Heading2', fontName='Roboto', fontSize=18, leading=22)
    styleSubH = ParagraphStyle(name='Heading3', fontName='Roboto', fontSize=14, leading=18)
    styleSmallH = ParagraphStyle(name='Heading4', fontName='Roboto', fontSize=12, leading=16)
    styleN = ParagraphStyle(name='Normal', fontName='Roboto', fontSize=11, leading=14)
    try:
        # Helper function
        def load_json_field(key):
            val = request.form.get(key, '')
            if not val:
                raise ValueError(f"Thiếu hoặc rỗng: {key}")
            return json.loads(val)
        # Lấy dữ liệu từ form hoặc query string
        data = load_json_field('data') or []
        criteria = load_json_field('criteria') or []
        weights = load_json_field('weights') or []
        pairwise_matrices = load_json_field('pairwise_matrices') or {}
        eigen_vectors = load_json_field('eigen_vectors') or {}
        vectonhatquan = load_json_field('vectonhatquan') or {}
        lambda_max_list = load_json_field('lambda_max_list') or []
        CI_list = load_json_field('CI_list') or []
        CR_list = load_json_field('CR_list') or []
        scores = load_json_field('scores') or []
        best_option = load_json_field('best_option')

        charts = load_json_field('charts') or {}

        # --- Phần tạo PDF giữ nguyên như trước ---
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30,leftMargin=30, topMargin=30,bottomMargin=18)
        elements = []
        styles = getSampleStyleSheet()
        styleH = ParagraphStyle(name='Heading2', fontName='Roboto', fontSize=18, leading=22)
        styleSubH = ParagraphStyle(name='Heading3', fontName='Roboto', fontSize=14, leading=18)
        styleSmallH = ParagraphStyle(name='Heading4', fontName='Roboto', fontSize=12, leading=16)
        styleN = ParagraphStyle(name='Normal', fontName='Roboto', fontSize=11, leading=14)


        elements.append(Paragraph("Báo cáo Phân tích Phương án (AHP)", styleH))
        elements.append(Spacer(1, 12))

        elements.append(Paragraph("1. Dữ liệu phương án:", styleSubH))
        data_table = [['Phương án'] + criteria]
        for i, row in enumerate(data):
            data_table.append([f"PA{i+1}"] + [f"{float(v):.4f}" for v in row])
        table = Table(data_table, hAlign='LEFT')
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Roboto'),  # 👈 Gán font Roboto cho toàn bộ bảng
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),
            ('GRID', (0,0), (-1,-1), 1, colors.black)
        ]))
        elements.append(table)
        elements.append(Spacer(1, 12))

        elements.append(Paragraph("2. Trọng số tiêu chí:", styleSubH))
        weights_table = [['Tiêu chí', 'Trọng số']]
        for c, w in zip(criteria, weights):
            elements.append(Paragraph(f"{c}: {float(w):.4f}", styleN))
        elements.append(Spacer(1, 12))

        elements.append(Paragraph("3. Kết quả tính toán chi tiết:", styleSmallH))
        for idx, crit in enumerate(criteria):
            elements.append(Paragraph(f"Tiêu chí: {crit}", styleSmallH))

            matrix = pairwise_matrices.get(crit, [])
            if matrix:
                mat_data = [[f"{crit}"] + [f"PA{i+1}" for i in range(3)]]
                for i in range(3):
                    mat_data.append([f"PA{i+1}"] + [f"{float(matrix[i][j]):.3f}" for j in range(3)])
                t = Table(mat_data)
                t.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), 'Roboto'),  # 👈 Gán font Roboto cho toàn bộ bảng
                    ('BACKGROUND', (0,0), (-1,0), colors.grey),
                    ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
                    ('ALIGN',(0,0),(-1,-1),'CENTER'),
                    ('GRID', (0,0), (-1,-1), 1, colors.black)
                ]))
                elements.append(t)
                elements.append(Spacer(1, 6))

            ev = eigen_vectors.get(crit, [])
            vnhq = vectonhatquan.get(crit, [])
            if ev and vnhq:
                ev_data = [['Phương án', 'Eigen Vector', 'Vector Nhất Quán']]
                for i in range(3):
                    ev_data.append([f"PA{i+1}", f"{float(ev[i]):.4f}", f"{float(vnhq[i]):.4f}"])
                t2 = Table(ev_data)
                t2.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), 'Roboto'), 
                    ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                    ('ALIGN',(0,0),(-1,-1),'CENTER'),
                    ('GRID', (0,0), (-1,-1), 1, colors.black)
                ]))
                elements.append(t2)
                elements.append(Spacer(1, 6))

            elements.append(Paragraph(f"λ max: {float(lambda_max_list[idx]):.4f}", styleN))
            elements.append(Paragraph(f"CI: {float(CI_list[idx]):.4f}", styleN))
            cr_val = float(CR_list[idx])
            cr_status = "✅ (Hợp lý)" if cr_val < 0.1 else "❌ (Không hợp lý)"
            elements.append(Paragraph(f"CR: {cr_val:.4f} {cr_status}", styleN))
            elements.append(Spacer(1, 12))

        elements.append(Paragraph("4. Điểm ưu tiên các phương án:", styleSmallH))
        scores_table = [['Phương án', 'Điểm ưu tiên']]
        for i, s in enumerate(scores):
            scores_table.append([f"PA{i+1}", f"{float(s):.4f}"])
        t_scores = Table(scores_table)
        t_scores.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Roboto'),
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),
            ('GRID', (0,0), (-1,-1), 1, colors.black)
        ]))
        elements.append(t_scores)
        elements.append(Spacer(1, 12))

        if best_option:
            elements.append(Paragraph(f"Phương án tốt nhất là: PA{int(best_option)}", styleSmallH))
            elements.append(Spacer(1, 12))

        for title, img_base64 in charts.items():
            elements.append(Paragraph(title.replace('_', ' ').capitalize(), styleSmallH))
            img_data = base64.b64decode(img_base64)
            img_buffer = io.BytesIO(img_data)
            img = Image(img_buffer, width=400, height=250)
            elements.append(img)
            elements.append(Spacer(1, 12))

        doc.build(elements)
        buffer.seek(0)

        return send_file(buffer, as_attachment=True, download_name="bao_cao_ahp.pdf", mimetype='application/pdf')

    except Exception as e:
        return f"Lỗi khi xuất PDF: {str(e)}"


@app.route('/ketqua', methods=['POST'])
def ketqua():
    import numpy as np
    from flask import request

    criteria = session.get('criteria', [])
    weights = session.get('weights', [])
    data = request.form.get('data_json')

    import json
    matrix = json.loads(data)

    # Ví dụ xử lý: chuẩn hóa và nhân trọng số
    norm = np.array(matrix) / np.linalg.norm(matrix, axis=0)
    weighted = norm * weights
    scores = weighted.sum(axis=1)
    best_index = int(np.argmax(scores))
    best_pa = f"PA{best_index + 1}"

    return render_template("ketqua.html", scores=scores, best_pa=best_pa, matrix=matrix, criteria=criteria)

def get_connection():
    return psycopg2.connect(
        host="localhost",
        database="AHP",
        user="postgres",
        password="123"
    )

def save_to_db(matrix_id, criteria_list, option1_list):
    try:
        conn = get_connection()
        cur = conn.cursor()

        for i in range(len(criteria_list)):
            row_criteria = criteria_list[i]
            col_criteria = "Phương án 1"
            value = float(option1_list[i]) if option1_list[i] else 0.0

            cur.execute("""
                INSERT INTO CriteriaComparison (matrix_id, criteria_row, criteria_col, value)
                VALUES (%s, %s, %s, %s)
            """, (matrix_id, row_criteria, col_criteria, value))

        conn.commit()
        cur.close()
        conn.close()

        return True
    except Exception as e:
        print(f"Lỗi khi lưu vào DB: {e}")
        return False

# @app.route('/criteria-matrix')
# def criteria_matrix():
#     matrix_id = request.args.get('matrix_id', default=None)

#     if not matrix_id:
#         return "Không có matrix_id được chọn!"

#     conn = get_connection()
#     cur = conn.cursor()
#     cur.execute("SELECT criteria_row, criteria_col, value FROM public.criteriacomparison WHERE matrix_id = %s", (matrix_id,))
#     rows = cur.fetchall()
#     conn.close()

#     criteria = []
#     matrix = {}

#     for row, col, val in rows:
#         if row not in criteria:
#             criteria.append(row)
#         if col not in criteria:
#             criteria.append(col)
#         if row not in matrix:
#             matrix[row] = {}
#         matrix[row][col] = val

#     return render_template("matrix_view.html", criteria=criteria, matrix=matrix, matrix_id=matrix_id)




@app.route('/criteria-matrix')
def criteria_matrix():
    matrix_id = request.args.get('matrix_id', default=None)

    if not matrix_id:
        return "Không có matrix_id được chọn!"

    conn = get_connection()
    cur = conn.cursor()

    # Truy vấn ma trận tiêu chí từ bảng criteriacomparison
    cur.execute("SELECT criteria_row, criteria_col, value FROM public.criteriacomparison WHERE matrix_id = %s", (matrix_id,))
    rows = cur.fetchall()

    criteria = []
    matrix = {}
    for row, col, val in rows:
        if row not in criteria:
            criteria.append(row)
        if col not in criteria:
            criteria.append(col)
        if row not in matrix:
            matrix[row] = {}
        matrix[row][col] = val

    # Truy vấn ma trận so sánh cặp và trọng số từ bảng pairwise_comparison
    cur.execute(
        "SELECT criterion, option1, option2, pairwise_value, weight FROM public.pairwise_comparison WHERE matrix_id = %s",
        (matrix_id,)
    )
    pairwise_rows = cur.fetchall()

    pairwise_matrices = {}
    eigen_vectors = {}
    for crit, opt1, opt2, val, weight in pairwise_rows:
        if crit not in pairwise_matrices:
            pairwise_matrices[crit] = np.zeros((3, 3))
        if crit not in eigen_vectors:
            eigen_vectors[crit] = np.zeros(3)

        if val is not None:
            i = int(opt1[-1]) - 1
            j = int(opt2[-1]) - 1
            pairwise_matrices[crit][i][j] = val
        if weight is not None:
            i = int(opt1[-1]) - 1
            eigen_vectors[crit][i] = weight

    # Truy vấn dữ liệu từ options_data dựa trên matrix_id
    cur.execute(
        "SELECT id, matrix_id, option_id, pin, performance, camera, screen, weight, connectivity, memory FROM public.options_data WHERE matrix_id = %s",
        (matrix_id,)
    )
    options_rows = cur.fetchall()

    # Chuyển đổi dữ liệu thành danh sách từ điển
    options_data = [
        {
            'id': row[0],
            'matrix_id': row[1],
            'option_id': row[2],
            'pin': row[3],
            'performance': row[4],
            'camera': row[5],
            'screen': row[6],
            'weight': row[7],
            'connectivity': row[8],
            'memory': row[9]
        } for row in options_rows
    ]

    conn.close()

    return render_template('matrix_view.html',
                          criteria=criteria,
                          matrix=matrix,
                          pairwise_matrices=pairwise_matrices,
                          eigen_vectors=eigen_vectors,
                          matrix_id=matrix_id,
                          options_data=options_data)



@app.route('/list-matrix')
def list_matrix():
    conn = get_connection()
    cur = conn.cursor()
    
    # Lấy danh sách tất cả matrix_id (không trùng)
    cur.execute("SELECT DISTINCT matrix_id FROM public.criteriacomparison ORDER BY matrix_id DESC")
    matrix_ids = [row[0] for row in cur.fetchall()]
    
    conn.close()
    
    return render_template('list_matrix.html', matrix_ids=matrix_ids)


@app.route('/matrix-history')
def matrix_history():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT DISTINCT matrix_id FROM public.criteriacomparison ORDER BY matrix_id DESC")
    rows = cur.fetchall()
    conn.close()

    # Trả về danh sách dưới dạng JSON
    matrix_ids = [row[0] for row in rows]
    return jsonify(matrix_ids)


# Route mới để lấy lịch sử các phương án
@app.route('/option-history')
def option_history():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT DISTINCT matrix_id FROM public.options_data ORDER BY matrix_id DESC")
    rows = cur.fetchall()
    conn.close()   

 # Trả về danh sách dưới dạng JSON
    options_ids = [row[0] for row in rows]
    return jsonify(options_ids)




# _--------------------------------------------AI Dự Đoán---------------------------------------
@app.route('/dudoanai', methods=['POST'])
def dudoanai():
# Kiểm tra xem có file được gửi lên không
    if 'file' not in request.files:
        return 'Không có file', 400
    
    file = request.files['file']
    
    # Đọc file Excel
    df = pd.read_excel(file)
    
    # Chuẩn hóa dữ liệu mới với scaler đã lưu
    scaled_new = scaler.transform(df)
    
    # Dự đoán PCA
    pca_result_new = pca.transform(scaled_new)
    
    # Tính điểm PCA và sắp xếp
    df['PCA_score'] = pca_result_new[:, 0]
    df_sorted = df.sort_values(by='PCA_score', ascending=False)
    
    # Lấy dòng tốt nhất
    best_option = df_sorted.head(1)
    
    # Lấy chỉ số dòng tốt nhất
    best_row_index = best_option.index[0] + 1  
    
    # Chuyển kết quả thành chuỗi để dễ dàng truyền vào template
    best_option_str = best_option.to_string(index=False)
    
    # Trả về phương án tốt nhất và chỉ số dòng
    return render_template('compare.html', result=best_option_str, best_row=best_row_index)



if __name__ == "__main__":
    app.run(debug=True)
