from flask import Flask, render_template, request, session, redirect, url_for, jsonify, send_file
import pandas as pd
import numpy as np
import os
import psycopg2
import datetime
import joblib
import io
import xlsxwriter
import json
import matplotlib
matplotlib.use('Agg')
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import seaborn as sns
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import base64

app = Flask(__name__)
app.secret_key = 'p123'
UPLOAD_FOLDER = 'Uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Tạo thư mục lưu biểu đồ nếu chưa có (dùng cục bộ, không cần trên Railway)
if not os.path.exists('static/images'):
    os.makedirs('static/images')

# Đăng ký font RobotoCondensed-Regular
font_dir = os.path.join('static', 'fonts')
font_path = os.path.join(font_dir, 'Roboto_Condensed-Regular.ttf')
if os.path.exists(font_path):
    pdfmetrics.registerFont(TTFont('Roboto_Condensed-Regular', font_path))

# Load mô hình
models_path = os.path.join(os.getcwd(), 'models')
scaler = joblib.load(os.path.join(models_path, 'minmax_scaler.pkl'))
pca = joblib.load(os.path.join(models_path, 'pca_model.pkl'))

# Cấu hình matplotlib
plt.rcParams['font.family'] = 'sans-serif'
plt.rcParams['font.sans-serif'] = ['Roboto_Condensed-Regular', 'Arial', 'DejaVu Sans']

def get_connection():
    return psycopg2.connect(os.getenv('DATABASE_URL'))

def generate_matrix_id():
    return f"matrix_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"

def ratio_to_ahp_scale(a, b):
    if a == b:
        return 1
    if a == 0 or b == 0:
        return 1
    ratio = a / b if a > b else b / a
    thresholds = [0.011, 1.4, 1.6, 1.8, 2.2, 2.5, 3, 4, 5, 6, 7, 8, 9]
    ahp_values = [1, 2, 3, 4, 5, 6, 7, 8, 9, 9, 9, 9, 9]
    for i, threshold in enumerate(thresholds):
        if ratio <= threshold:
            value = ahp_values[i]
            break
    else:
        value = 9
    return value if a > b else 1 / value

def save_to_postgres(data, weights):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM options_data")
    matrix_id = generate_matrix_id()
    for i, row in enumerate(data, 1):
        cur.execute(
            "INSERT INTO options_data (option_id, pin, performance, camera, screen, weight, connectivity, memory, matrix_id) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)",
            (f'PA{i}', row[0], row[1], row[2], row[3], row[4], row[5], row[6], matrix_id)
        )
    cur.execute(
        "INSERT INTO options_data (option_id, pin, performance, camera, screen, weight, connectivity, memory, matrix_id) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)",
        ('Weights', weights[0], weights[1], weights[2], weights[3], weights[4], weights[5], weights[6], matrix_id)
    )
    conn.commit()
    cur.close()
    conn.close()

def save_pairwise_to_postgres(criterion, matrix, weights):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM pairwise_comparison WHERE criterion = %s", (criterion,))
    matrix_id = generate_matrix_id()
    for i in range(3):
        for j in range(3):
            cur.execute(
                "INSERT INTO pairwise_comparison (criterion, option1, option2, pairwise_value, weight, matrix_id) VALUES (%s, %s, %s, %s, %s, %s)",
                (criterion, f'PA{i+1}', f'PA{j+1}', float(matrix[i][j]), None, matrix_id)
            )
    for i, weight in enumerate(weights, 1):
        cur.execute(
            "INSERT INTO pairwise_comparison (criterion, option1, option2, pairwise_value, weight, matrix_id) VALUES (%s, %s, %s, %s, %s, %s)",
            (criterion, f'PA{i}', 'Weight', None, float(weight), matrix_id)
        )
    conn.commit()
    cur.close()
    conn.close()

def save_scores(scores):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM option_scores")
    matrix_id = generate_matrix_id()
    for i, score in enumerate(scores, 1):
        cur.execute(
            "INSERT INTO option_scores (option_id, score, matrix_id) VALUES (%s, %s, %s)",
            (f'PA{i}', float(score), matrix_id)
        )
    conn.commit()
    cur.close()
    conn.close()

@app.route("/", methods=["GET", "POST"])
def index():
    return render_template("index.html")

@app.route("/compare", methods=["GET", "POST"])
def compare():
    return render_template("compare.html")

@app.route('/result', methods=['POST'])
def result():
    matrix_id = generate_matrix_id()
    criteria = request.form.getlist('criteria[]')
    options = []
    index = 1
    while True:
        key = f'option{index}[]'
        if key in request.form:
            values = [float(x) if x else 0.0 for x in request.form.getlist(key)]
            options.append(values)
            index += 1
        else:
            break

    data_matrix = np.array(options).T
    col_sums = np.sum(data_matrix, axis=0)
    normalized_data = np.divide(data_matrix, col_sums, where=col_sums != 0)
    rounded_data = np.round(normalized_data, 4)
    weights = np.mean(rounded_data, axis=1)
    weights = np.round(weights, 4).tolist()
    Aw = np.dot(data_matrix, weights)
    row_sums = np.sum(np.array([data_matrix[:, j] * weights[j] for j in range(len(weights))]), axis=0)
    consistency_vector = np.divide(row_sums, weights, where=np.array(weights) != 0)
    lambda_max = np.mean(consistency_vector)
    n = len(criteria)
    CI = (lambda_max - n) / (n - 1) if n > 1 else 0
    RI_dict = {1: 0.0, 2: 0.0, 3: 0.58, 4: 0.90, 5: 1.12, 6: 1.24, 7: 1.32, 8: 1.41, 9: 1.45, 10: 1.49}
    RI = RI_dict.get(n, 1.49)
    CR = CI / RI if RI != 0 else 0

    # Tạo biểu đồ
    charts = {}
    # 1. Biểu đồ cột
    plt.figure(figsize=(8, 5))
    plt.bar(criteria, weights, color='skyblue')
    plt.title('Trọng số của từng tiêu chí')
    plt.xlabel('Tiêu chí')
    plt.ylabel('Trọng số')
    plt.grid(axis='y', linestyle='--', alpha=0.7)
    img = io.BytesIO()
    plt.savefig(img, format='png')
    img.seek(0)
    charts['bar_chart'] = base64.b64encode(img.getvalue()).decode()
    img.close()
    plt.close()

    # 2. Biểu đồ radar
    angles = np.linspace(0, 2 * np.pi, len(criteria), endpoint=False).tolist()
    weights_plot = weights + weights[:1]
    angles += angles[:1]
    fig, ax = plt.subplots(figsize=(6, 6), subplot_kw=dict(polar=True))
    ax.fill(angles, weights_plot, color='skyblue', alpha=0.25)
    ax.plot(angles, weights_plot, color='blue', linewidth=2)
    ax.set_yticklabels([])
    ax.set_xticks(angles[:-1])
    ax.set_xticklabels(criteria)
    plt.title('Biểu đồ Radar - Trọng số tiêu chí')
    img = io.BytesIO()
    plt.savefig(img, format='png')
    img.seek(0)
    charts['radar_chart'] = base64.b64encode(img.getvalue()).decode()
    img.close()
    plt.close()

    # 3. Biểu đồ nhiệt
    plt.figure(figsize=(8, 6))
    sns.heatmap(data_matrix, annot=True, fmt='.2f', cmap='Blues', xticklabels=criteria, yticklabels=criteria)
    plt.title('Ma trận so sánh đôi')
    img = io.BytesIO()
    plt.savefig(img, format='png')
    img.seek(0)
    charts['heatmap'] = base64.b64encode(img.getvalue()).decode()
    img.close()
    plt.close()

    # Lưu vào CSDL
    try:
        conn = get_connection()
        cur = conn.cursor()
        for j, col in enumerate(options):
            for i, val in enumerate(col):
                cur.execute(
                    "INSERT INTO CriteriaComparison (matrix_id, criteria_row, criteria_col, value) VALUES (%s, %s, %s, %s)",
                    (matrix_id, criteria[i], criteria[j], val)
                )
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        return f"❌ Lỗi khi lưu vào CSDL: {str(e)}"

    session['criteria'] = criteria
    session['weights'] = weights
    session['charts'] = charts

    return render_template(
        'result.html',
        criteria=criteria,
        data=rounded_data.tolist(),
        alternatives=criteria,
        weights=weights,
        aw=Aw.tolist(),
        data_matrix=data_matrix.tolist(),
        weighted_matrix=np.array([data_matrix[:, j] * weights[j] for j in range(len(weights))]).T.tolist(),
        row_sums=row_sums.tolist(),
        consistency_vector=consistency_vector.tolist(),
        lambda_max=lambda_max,
        ci=CI,
        cr=CR,
        weights_vector=weights,
        charts=charts
    )

@app.route('/luachontieuchi', methods=['GET', 'POST'])
def luachontieuchi():
    criteria = session.get('criteria', [])
    weights = session.get('weights', [])
    return render_template('luachontieuchi.html', criteria=criteria, weights=weights)

@app.route('/combined_input', methods=['GET', 'POST'])
def combined_input():
    criteria = session.get('criteria', [])
    weights = session.get('weights', [])
    data = session.get('data', [])
    message = ""
    scores = session.get('scores', [])
    best_option = session.get('best_option', None)
    pairwise_matrices = session.get('pairwise_matrices', {})
    eigen_vectors = session.get('eigen_vectors', {})
    vectonhatquan = session.get('vectonhatquan', {})
    lambda_max_list = session.get('lambda_max_list', [])
    CI_list = session.get('CI_list', [])
    CR_list = session.get('CR_list', [])
    charts = session.get('charts', {})
    RI_dict = {1: 0.00, 2: 0.00, 3: 0.58, 4: 0.90, 5: 1.12, 6: 1.24, 7: 1.32, 8: 1.41, 9: 1.45, 10: 1.49}

    if request.method == 'POST':
        keys_to_clear = ['data', 'pairwise_matrices', 'eigen_vectors', 'vectonhatquan', 'lambda_max_list', 'CI_list', 'CR_list', 'scores', 'best_option', 'charts']
        for key in keys_to_clear:
            session.pop(key, None)

        file = request.files.get('excel_file')
        if file and file.filename.endswith('.xlsx'):
            try:
                df = pd.read_excel(file)
                if len(df.columns) >= len(criteria):
                    data = df.iloc[:3, :len(criteria)].values.tolist()
                    message = "Đọc file Excel thành công!"
                else:
                    message = "Số cột trong file Excel không đủ!"
            except Exception as e:
                message = f"Lỗi khi đọc file Excel: {e}"
        else:
            data = []
            for i in range(3):
                row = request.form.getlist(f'option{i+1}[]')
                if row:
                    try:
                        data.append([float(x) if x else 0.0 for x in row])
                    except:
                        data.append([0.0 for _ in range(len(criteria))])
            message = "Dữ liệu nhập tay đã được xử lý!"

        if data and weights and len(data[0]) == len(weights):
            save_to_postgres(data, weights)
            pairwise_matrices = {}
            eigen_vectors = {}
            vectonhatquan = {}
            lambda_max_list = []
            CI_list = []
            CR_list = []

            for j, crit in enumerate(criteria):
                col = np.array([row[j] for row in data])
                matrix = np.zeros((3, 3))
                for i in range(3):
                    for k in range(3):
                        matrix[i][k] = ratio_to_ahp_scale(col[i], col[k])
                pairwise_matrices[crit] = matrix
                eigen_vector = np.mean(matrix / np.sum(matrix, axis=0), axis=1)
                save_pairwise_to_postgres(crit, matrix, eigen_vector.round(4))
                column_sums = np.sum(matrix, axis=0)
                normalized_matrix = matrix / column_sums
                eigen_vector = np.mean(normalized_matrix, axis=1)
                eigen_vectors[crit] = eigen_vector.round(4)
                consistency_vector = np.dot(matrix, eigen_vector) / eigen_vector
                vectonhatquan[crit] = consistency_vector.round(4)
                lambda_max = np.sum(np.dot(matrix, eigen_vector) / eigen_vector) / 3
                CI = (lambda_max - 3) / (3 - 1)
                RI = RI_dict[3]
                CR = CI / RI if RI != 0 else 0
                lambda_max_list.append(round(lambda_max, 4))
                CI_list.append(round(CI, 4))
                CR_list.append(round(CR, 4))

            scores = []
            for row in data:
                score = sum([v * w for v, w in zip(row, weights)])
                scores.append(round(score, 4))
            best_option = scores.index(max(scores)) + 1
            save_scores(scores)

            charts = {}
            # 1. Biểu đồ cột cho trọng số tiêu chí
            plt.figure(figsize=(8, 5))
            plt.bar(criteria, weights, color='skyblue')
            plt.title('Trọng số của các tiêu chí')
            plt.xlabel('Tiêu chí')
            plt.ylabel('Trọng số')
            plt.grid(axis='y', linestyle='--', alpha=0.7)
            img = io.BytesIO()
            plt.savefig(img, format='png')
            img.seek(0)
            charts['criteria_weights_bar'] = base64.b64encode(img.getvalue()).decode()
            img.close()
            plt.close()

            # 2. Biểu đồ cột cho điểm ưu tiên
            plt.figure(figsize=(6, 4))
            plt.bar(['PA1', 'PA2', 'PA3'], scores, color='lightgreen')
            plt.title('Điểm ưu tiên của các phương án')
            plt.xlabel('Phương án')
            plt.ylabel('Điểm')
            plt.grid(axis='y', linestyle='--', alpha=0.7)
            img = io.BytesIO()
            plt.savefig(img, format='png')
            img.seek(0)
            charts['scores_bar'] = base64.b64encode(img.getvalue()).decode()
            img.close()
            plt.close()

            # 3. Biểu đồ radar cho điểm số phương án
            angles = np.linspace(0, 2 * np.pi, len(criteria), endpoint=False).tolist()
            angles += angles[:1]
            fig, ax = plt.subplots(figsize=(6, 6), subplot_kw=dict(polar=True))
            for i, row in enumerate(data):
                values = row + row[:1]
                ax.plot(angles, values, linewidth=2, label=f'PA{i+1}')
                ax.fill(angles, values, alpha=0.25)
            ax.set_xticks(angles[:-1])
            ax.set_xticklabels(criteria)
            ax.set_yticklabels([])
            plt.title('Điểm số của phương án theo tiêu chí')
            plt.legend(loc='upper right')
            img = io.BytesIO()
            plt.savefig(img, format='png')
            img.seek(0)
            charts['options_radar'] = base64.b64encode(img.getvalue()).decode()
            img.close()
            plt.close()

            # 4. Biểu đồ nhiệt cho ma trận so sánh cặp
            plt.figure(figsize=(6, 5))
            sns.heatmap(pairwise_matrices[criteria[0]], annot=True, fmt='.2f', cmap='Blues', xticklabels=['PA1', 'PA2', 'PA3'], yticklabels=['PA1', 'PA2', 'PA3'])
            plt.title(f'Ma trận so sánh cặp - Tiêu chí: {criteria[0]}')
            img = io.BytesIO()
            plt.savefig(img, format='png')
            img.seek(0)
            charts['pairwise_heatmap_first'] = base64.b64encode(img.getvalue()).decode()
            img.close()
            plt.close()

            # 5. Biểu đồ cột nhóm cho trọng số phương án
            fig, ax = plt.subplots(figsize=(10, 6))
            bar_width = 0.25
            x = np.arange(len(criteria))
            for i in range(3):
                eigen_values = [eigen_vectors[crit][i] for crit in criteria]
                plt.bar(x + i * bar_width, eigen_values, bar_width, label=f'PA{i+1}')
            plt.xlabel('Tiêu chí')
            plt.ylabel('Trọng số')
            plt.title('Trọng số của phương án theo tiêu chí')
            plt.xticks(x + bar_width, criteria)
            plt.legend()
            plt.grid(axis='y', linestyle='--', alpha=0.7)
            img = io.BytesIO()
            plt.savefig(img, format='png')
            img.seek(0)
            charts['eigen_vectors_bar'] = base64.b64encode(img.getvalue()).decode()
            img.close()
            plt.close()

            session['data'] = data
            session['pairwise_matrices'] = {k: v.tolist() for k, v in pairwise_matrices.items()}
            session['eigen_vectors'] = {k: v.tolist() for k, v in eigen_vectors.items()}
            session['vectonhatquan'] = {k: v.tolist() for k, v in vectonhatquan.items()}
            session['lambda_max_list'] = lambda_max_list
            session['CI_list'] = CI_list
            session['CR_list'] = CR_list
            session['scores'] = scores
            session['best_option'] = best_option
            session['charts'] = charts

    return render_template('combined_input.html',
                          criteria=criteria,
                          weights=weights,
                          data=data,
                          message=message,
                          scores=scores,
                          best_option=best_option,
                          pairwise_matrices=pairwise_matrices,
                          eigen_vectors=eigen_vectors,
                          vectonhatquan=vectonhatquan,
                          lambda_max_list=lambda_max_list,
                          CI_list=CI_list,
                          CR_list=CR_list,
                          charts=charts)

@app.route('/export_excel')
def export_excel():
    criteria = session.get('criteria', [])
    weights = session.get('weights', [])
    data = session.get('data', [])
    pairwise_matrices = session.get('pairwise_matrices', {})
    eigen_vectors = session.get('eigen_vectors', {})
    lambda_max_list = session.get('lambda_max_list', [])
    CI_list = session.get('CI_list', [])
    CR_list = session.get('CR_list', [])
    scores = session.get('scores', [])
    best_option = session.get('best_option', 1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Kết Quả AHP"
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    ws.append(["Bảng Dữ Liệu Nhập Tay"])
    ws.append(["Phương án"] + criteria)
    for i, row in enumerate(data, start=1):
        ws.append([f"PA{i}"] + [round(val, 4) for val in row])
    ws.append(["Trọng số"] + [round(w, 4) for w in weights])
    ws.append([])

    for row in ws[1:len(data)+3]:
        for cell in row:
            cell.font = bold_font if cell.row == 1 or cell.row == len(data)+2 else Font()
            cell.alignment = center_alignment
            if cell.row == 1:
                cell.fill = header_fill

    row_offset = len(data) + 4
    for idx, (crit, matrix) in enumerate(pairwise_matrices.items()):
        ws.append([f"Ma trận so sánh cặp - Tiêu chí: {crit}"])
        ws.append(["", "PA1", "PA2", "PA3", "Trọng số"])
        for i in range(3):
            row_data = [f"PA{i+1}"] + [round(val, 3) for val in matrix[i]] + [round(eigen_vectors[crit][i], 3)]
            ws.append(row_data)
        ws.append([f"λ max: {round(lambda_max_list[idx], 4)}"])
        ws.append([f"CI: {round(CI_list[idx], 4)}"])
        ws.append([f"CR: {round(CR_list[idx], 4)} {'(Hợp lý)' if CR_list[idx] < 0.1 else '(Không hợp lý)'}"])
        ws.append([])
        row_offset += 7
        for row in ws[row_offset:row_offset+4]:
            for cell in row:
                cell.font = bold_font if cell.row == row_offset or cell.column == 6 else Font()
                cell.alignment = center_alignment
                if cell.row == row_offset:
                    cell.fill = header_fill
        row_offset += 4

    ws.append(["Kết quả tính độ ưu tiên"])
    ws.append(["Phương án", "Điểm ưu tiên"])
    for i, score in enumerate(scores, start=1):
        ws.append([f"PA{i}", round(score, 4)])
    ws.append([f"Phương án tốt nhất: PA{best_option}"])
    ws.append([])

    for row in ws[row_offset:row_offset+len(scores)+2]:
        for cell in row:
            cell.font = bold_font if cell.row == row_offset or cell.row == row_offset+len(scores)+1 else Font()
            cell.alignment = center_alignment
            if cell.row == row_offset:
                cell.fill = header_fill

    for col in range(1, ws.max_column+1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name="Ket_Qua_AHP.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route('/export_pdf')
def export_pdf():
    criteria = session.get('criteria', [])
    weights = session.get('weights', [])
    scores = session.get('scores', [])
    best_option = session.get('best_option')
    pairwise_matrices = session.get('pairwise_matrices', {})
    eigen_vectors = session.get('eigen_vectors', {})
    lambda_max_list = session.get('lambda_max_list', [])
    CI_list = session.get('CI_list', [])
    CR_list = session.get('CR_list', [])
    charts = session.get('charts', {})

    if not scores:
        return "Không có dữ liệu để xuất PDF", 400

    pdf_io = io.BytesIO()
    doc = SimpleDocTemplate(pdf_io, pagesize=A4)
    elements = []

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='TitleVN', fontName='Roboto_Condensed-Regular', fontSize=16, leading=20, alignment=1))
    styles.add(ParagraphStyle(name='Heading2VN', fontName='Roboto_Condensed-Regular', fontSize=14, leading=16))
    styles.add(ParagraphStyle(name='Heading3VN', fontName='Roboto_Condensed-Regular', fontSize=12, leading=14))
    styles.add(ParagraphStyle(name='NormalVN', fontName='Roboto_Condensed-Regular', fontSize=10, leading=12))

    elements.append(Paragraph("Báo cáo Kết quả Phân tích AHP", styles['TitleVN']))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("Kết quả tính độ ưu tiên", styles['Heading2VN']))
    scores_data = [['Phương án', 'Điểm ưu tiên']] + [[f'PA{i+1}', score] for i, score in enumerate(scores)]
    scores_table = Table(scores_data)
    scores_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Roboto_Condensed-Regular'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(scores_table)
    elements.append(Spacer(1, 12))

    elements.append(Paragraph(f"Phương án tốt nhất: PA{best_option}", styles['NormalVN']))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("Ma trận so sánh cặp cho từng tiêu chí", styles['Heading2VN']))
    for idx, (crit, matrix) in enumerate(pairwise_matrices.items()):
        elements.append(Paragraph(f"Tiêu chí: {crit}", styles['Heading3VN']))
        matrix_data = [['', 'PA1', 'PA2', 'PA3', 'Trọng số']] + \
                     [[f'PA{i+1}'] + [f'{val:.3f}' for val in row] + [f'{eigen_vectors[crit][i]:.3f}'] for i, row in enumerate(matrix)]
        matrix_table = Table(matrix_data)
        matrix_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Roboto_Condensed-Regular'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(matrix_table)
        elements.append(Paragraph(f"λ max: {lambda_max_list[idx]}", styles['NormalVN']))
        elements.append(Paragraph(f"CI: {CI_list[idx]}", styles['NormalVN']))
        elements.append(Paragraph(f"CR: {CR_list[idx]} {'✅ (Hợp lý)' if CR_list[idx] < 0.1 else '❌ (Không hợp lý)'}", styles['NormalVN']))
        elements.append(Spacer(1, 12))

    doc.build(elements)
    pdf_io.seek(0)

    return send_file(pdf_io,
                     download_name='ahp_report.pdf',
                     as_attachment=True,
                     mimetype='application/pdf')

@app.route('/ketqua', methods=['POST'])
def ketqua():
    criteria = session.get('criteria', [])
    weights = session.get('weights', [])
    data = request.form.get('data_json')
    matrix = json.loads(data)
    norm = np.array(matrix) / np.linalg.norm(matrix, axis=0)
    weighted = norm * weights
    scores = weighted.sum(axis=1)
    best_index = int(np.argmax(scores))
    best_pa = f"PA{best_index + 1}"
    return render_template("ketqua.html", scores=scores, best_pa=best_pa, matrix=matrix, criteria=criteria)

def save_to_db(matrix_id, criteria_list, option1_list):
    try:
        conn = get_connection()
        cur = conn.cursor()
        for i in range(len(criteria_list)):
            row_criteria = criteria_list[i]
            col_criteria = "Phương án 1"
            value = float(option1_list[i]) if option1_list[i] else 0.0
            cur.execute(
                "INSERT INTO CriteriaComparison (matrix_id, criteria_row, criteria_col, value) VALUES (%s, %s, %s, %s)",
                (matrix_id, row_criteria, col_criteria, value)
            )
        conn.commit()
        cur.close()
        conn.close()
        return True
    except Exception as e:
        print(f"Lỗi khi lưu vào DB: {e}")
        return False

@app.route('/criteria-matrix')
def criteria_matrix():
    matrix_id = request.args.get('matrix_id', default=None)
    if not matrix_id:
        return "Không có matrix_id được chọn!"
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT criteria_row, criteria_col, value FROM public.criteriacomparison WHERE matrix_id = %s", (matrix_id,))
    rows = cur.fetchall()
    criteria = []
    matrix = {}
    for row, col, val in rows:
        if row not in criteria:
            criteria.append(row)
        if col not in criteria:
            criteria.append(col)
        if row not in matrix:
            matrix[row] = {}
        matrix[row][col] = val

    cur.execute(
        "SELECT criterion, option1, option2, pairwise_value, weight FROM public.pairwise_comparison WHERE matrix_id = %s",
        (matrix_id,)
    )
    pairwise_rows = cur.fetchall()
    pairwise_matrices = {}
    eigen_vectors = {}
    for crit, opt1, opt2, val, weight in pairwise_rows:
        if crit not in pairwise_matrices:
            pairwise_matrices[crit] = np.zeros((3, 3))
        if crit not in eigen_vectors:
            eigen_vectors[crit] = np.zeros(3)
        if val is not None:
            i = int(opt1[-1]) - 1
            j = int(opt2[-1]) - 1
            pairwise_matrices[crit][i][j] = val
        if weight is not None:
            i = int(opt1[-1]) - 1
            eigen_vectors[crit][i] = weight

    cur.execute(
        "SELECT id, matrix_id, option_id, pin, performance, camera, screen, weight, connectivity, memory FROM public.options_data WHERE matrix_id = %s",
        (matrix_id,)
    )
    options_rows = cur.fetchall()
    options_data = [
        {
            'id': row[0],
            'matrix_id': row[1],
            'option_id': row[2],
            'pin': row[3],
            'performance': row[4],
            'camera': row[5],
            'screen': row[6],
            'weight': row[7],
            'connectivity': row[8],
            'memory': row[9]
        } for row in options_rows
    ]
    conn.close()
    return render_template('matrix_view.html',
                          criteria=criteria,
                          matrix=matrix,
                          pairwise_matrices=pairwise_matrices,
                          eigen_vectors=eigen_vectors,
                          matrix_id=matrix_id,
                          options_data=options_data)

@app.route('/list-matrix')
def list_matrix():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT DISTINCT matrix_id FROM public.criteriacomparison ORDER BY matrix_id DESC")
    matrix_ids = [row[0] for row in cur.fetchall()]
    conn.close()
    return render_template('list_matrix.html', matrix_ids=matrix_ids)

@app.route('/matrix-history')
def matrix_history():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT DISTINCT matrix_id FROM public.criteriacomparison ORDER BY matrix_id DESC")
    rows = cur.fetchall()
    conn.close()
    matrix_ids = [row[0] for row in rows]
    return jsonify(matrix_ids)

@app.route('/option-history')
def option_history():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT DISTINCT matrix_id FROM public.options_data ORDER BY matrix_id DESC")
    rows = cur.fetchall()
    conn.close()
    options_ids = [row[0] for row in rows]
    return jsonify(options_ids)

@app.route('/dudoanai', methods=['POST'])
def dudoanai():
    if 'file' not in request.files:
        return 'Không có file', 400
    file = request.files['file']
    df = pd.read_excel(file)
    scaled_new = scaler.transform(df)
    pca_result_new = pca.transform(scaled_new)
    df['PCA_score'] = pca_result_new[:, 0]
    df_sorted = df.sort_values(by='PCA_score', ascending=False)
    best_option = df_sorted.head(1)
    best_row_index = best_option.index[0] + 1
    best_option_str = best_option.to_string(index=False)
    return render_template('compare.html', result=best_option_str, best_row=best_row_index)

if __name__ == "__main__":
    app.run(debug=True)